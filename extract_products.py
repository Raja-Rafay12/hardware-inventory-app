import sys
import os
import openpyxl
import json

def categorize(name):
    name_lower = name.lower()
    
    # Electrical keywords
    elec_kws = ["wire", "cable", "switch", "socket", "tester", "holder", "smd", "bulb", "breaker", 
                "db ", "meter box", "tape osaka", "insulation tape", "thimmal", "piano", "capacitor", "saddle 25mm", "saddle 32mm"]
    # Plumbing & Sanitary keywords
    plumb_kws = ["pipe", "nipple", "elbow", "tee", "socket", "union", "valve", "cock", "shower", 
                 "bason", "basin", "flush", "waste", "jali", "tanki", "ppr", "pvc", "gi ", "bush", "connection pipe", "mixer pipe", "nrv", "saddle"]
    # Hardware & Tools keywords
    tool_kws = ["disc", "cutting", "grinding", "steel", "welding", "nail", "keel", "hinges", "qabza", 
                "screw", "bolt", "ficher", "lock", "wahoo", "wahu", "key", "hammer", "wrench", "pliers", 
                "cutter", "gainti", "bailcha", "chisel", "saw", "tape measuring", "measuring tape", "tape masking", "masking tape", 
                "gloves", "glove", "brush", "dhaga", "sootar", "level", "karandi", "fara", "bracket", "clamp", "unifix", "doori"]
    # Paint keywords
    paint_kws = ["paint", "oil paint", "enamel", "roller", "varnish", "colour", "putty", "solution", 
                 "samad", "elfi", "glue"]
    # Cement & Aggregates keywords
    cement_kws = ["sand paper", "cement", "bond"]
    
    if any(kw in name_lower for kw in elec_kws):
        return "Electrical"
    if any(kw in name_lower for kw in plumb_kws):
        return "Plumbing & Sanitary"
    if any(kw in name_lower for kw in paint_kws):
        return "Paint"
    if any(kw in name_lower for kw in cement_kws):
        return "Cement & Aggregates"
    if any(kw in name_lower for kw in tool_kws):
        return "Hardware & Tools"
        
    return "General"

def main():
    user_home = os.path.expanduser('~')
    excel_path = os.path.join(user_home, 'Documents', 'NEW PRICE LIST 2024.xlsx')
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # We will read both 'PRICES' and 'PRICES (2)' sheets just in case. 
    # Usually they are duplicate or one is more complete. Let's merge them.
    sheets_to_read = ['PRICES', 'PRICES (2)']
    products_dict = {} # name -> product_dict
    
    for sheet_name in sheets_to_read:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found, skipping...")
            continue
            
        sheet = wb[sheet_name]
        print(f"Processing sheet '{sheet_name}'...")
        row_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True): # Skip header
            row_count += 1
            if len(row) < 7:
                continue
                
            item_name = row[1]
            if not item_name:
                continue
                
            item_name = str(item_name).strip()
            if not item_name or item_name == "0" or item_name.startswith("=="):
                continue
                
            # Quantity
            qty_val = row[2]
            try:
                quantity = int(float(qty_val)) if qty_val is not None else 0
            except ValueError:
                quantity = 0
                
            # Unit
            unit_val = row[3]
            unit = str(unit_val).strip().lower() if unit_val else "piece"
            if unit in ["", "none"]:
                unit = "piece"
            elif unit in ["pkt", "packet"]:
                unit = "packet"
            elif unit in ["kg"]:
                unit = "kg"
            elif unit in ["pair"]:
                unit = "pair"
                
            # Cost Price
            cost_val = row[4]
            try:
                cost_price = float(cost_val) if cost_val is not None else 0.0
            except ValueError:
                cost_price = 0.0
                
            # Sale Price
            sale_val = row[5]
            try:
                sale_price = float(sale_val) if sale_val is not None else 0.0
            except ValueError:
                sale_price = 0.0
                
            # Calculate Markup
            markup = 40.0
            if cost_price > 0 and sale_price > 0:
                markup = ((sale_price - cost_price) / cost_price) * 100.0
                markup = round(markup, 1)
                if markup < 0 or markup > 1000:
                    markup = 40.0
                    
            category = categorize(item_name)
            
            # Save or update product (since it's ordered chronologically, we overwrite with later rows)
            products_dict[item_name] = {
                "name": item_name,
                "category": category,
                "unit": unit,
                "quantity": 0, # Starting quantity in store is 0 or what's in excel? Let's make it 0 as it is seed inventory.
                "costPrice": round(cost_price, 2),
                "markup": markup,
                "lowStock": 5
            }
            
        print(f"  Read {row_count} rows. Total unique items so far: {len(products_dict)}")
        
    # Sort products alphabetically by name
    sorted_items = sorted(products_dict.values(), key=lambda x: x["name"])
    
    # Assign IDs
    final_products = []
    for idx, item in enumerate(sorted_items):
        item["id"] = f"p{idx+1:04d}"
        final_products.append(item)
        
    print(f"\nFinal extracted unique products count: {len(final_products)}")
    
    # Print stats by category
    cat_counts = {}
    for item in final_products:
        cat = item["category"]
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
    print("Category breakdown:")
    for cat, count in sorted(cat_counts.items(), key=lambda x: x[1], reverse=True):
        print(f"  - {cat}: {count}")
        
    # Save output
    output_path = os.path.join(os.path.dirname(__file__), 'products.json')
    with open(output_path, 'w') as f:
        json.dump(final_products, f, indent=2)
    print(f"Saved products JSON to {output_path}")

if __name__ == "__main__":
    main()
