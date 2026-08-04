import sys
import os
import openpyxl

def main():
    user_home = os.path.expanduser('~')
    file_path = os.path.join(user_home, 'Documents', 'NEW PRICE LIST 2024.xlsx')
    if not os.path.exists(file_path):
        print(f"Error: file not found at {file_path}")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Worksheets available in the workbook:")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        dim = getattr(sheet, 'dimensions', 'Unknown')
        print(f"  - Sheet: '{sheet_name}', Dimensions: {dim}")
        
        # Read the first few rows to see columns
        rows = list(sheet.iter_rows(max_row=5, values_only=True))
        if rows:
            print("    First few rows:")
            for i, r in enumerate(rows):
                # Print row elements, filtering out None or truncation
                print(f"      Row {i+1}: {r[:10]}")
        else:
            print("    Sheet is empty")

if __name__ == "__main__":
    main()
